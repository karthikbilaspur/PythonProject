import os
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd

def reader(file, path):
    abs_file = os.path.join(path, file)
    wb_sheet = load_workbook(abs_file).active
    rows = []
    # min_row is set to 2, ignore the first row which contains headers
    for row in wb_sheet.iter_rows(min_row=2):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        rows.append(row_data)
    return rows

def merge_excel_files(path, output_file):
    book = Workbook()
    sheet = book.active
    files = os.listdir(path)
    for file in files:
        rows = reader(file, path)
        for row in rows:
            sheet.append(row)
    book.save(filename=output_file)

def merge_excel_files_pandas(file_paths, output_file):
    dfs = []
    for file_path in file_paths:
        df = pd.read_excel(file_path)
        dfs.append(df)
    merged_df = pd.concat(dfs, ignore_index=True)
    merged_df.to_excel(output_file, index=False)

# Example usage
if __name__ == '__main__':
    path = input('Path: ')
    output_file = input('Unified Workbook name ')
    merge_excel_files(path, output_file)
    # or
    file_paths = [os.path.join(path, file) for file in os.listdir(path)]
    merge_excel_files_pandas(file_paths, output_file)